import io
from calendar import monthrange
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

TURKISH_DAYS = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]
TURKISH_MONTHS = [
    "", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]

# ── Fills / fonts ──────────────────────────────────────────────────────────────
_H1   = PatternFill("solid", fgColor="1F4E79")   # dark navy  – header row 1
_H2   = PatternFill("solid", fgColor="2E75B6")   # medium blue – header row 2
_TOT  = PatternFill("solid", fgColor="C00000")   # dark red   – monthly total
_WKD  = PatternFill("solid", fgColor="EBEBEB")   # light grey  – weekend
_HOL  = PatternFill("solid", fgColor="FFF2CC")   # light yellow – holiday
_RED  = PatternFill("solid", fgColor="FF4444")   # red  – negative FARK
_GRN  = PatternFill("solid", fgColor="70AD47")   # green – positive FARK

_WB   = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_THIN = Side(style="thin")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RGT  = Alignment(horizontal="right",  vertical="center")
_LFT  = Alignment(horizontal="left",   vertical="center")
_MNY  = '"₺"#,##0.00'


def _c(ws, row, col, value=None, fill=None, font=None, fmt=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:  c.fill  = fill
    if font:  c.font  = font
    if fmt:   c.number_format = fmt
    c.alignment = align or _CTR
    c.border    = _BRD
    return c


# ── Column layout ─────────────────────────────────────────────────────────────
#  A=1  TARİH
#  B=2  ÇIKAN
#  C=3  TEB             ] BANKA (POS)
#  D=4  VAKİFBANK       ]
#  E=5  NAKİT (GİRİŞ)   ] NAKİT
#  F=6  İBAN            ]
#  G=7  İADE BANKA      ] İADE
#  H=8  İADE NAKİT      ]
#  I=9  TAHSİLAT BANKA  ] TAHSİLAT
#  J=10 TAHSİLAT NAKİT  ]
#  K=11 BANKA TOPLAM
#  L=12 NAKİT TOPLAM
#  M=13 GENEL TOPLAM
#  N=14 RX BANKA        ] RX
#  O=15 RX NAKİT        ]
#  P=16 RX TOPLAM       ]
#  Q=17 FARK BANKA      ] FARK
#  R=18 FARK NAKİT      ]
#  S=19 TOPLAM FARK     ]
#  T=20 NOT

def _calc(e):
    cikan          = float(e.get("cikan")          or 0)
    teb            = float(e.get("teb")            or 0)
    vakifbank      = float(e.get("vakifbank")      or 0)
    nakit          = float(e.get("nakit")          or 0)
    iban           = float(e.get("iban")           or 0)
    iade_nakit     = float(e.get("iade_nakit")     or 0)
    iade_banka     = float(e.get("iade_banka")     or 0)
    tahsilat_banka = float(e.get("tahsilat_banka") or 0)
    tahsilat_nakit = float(e.get("tahsilat_nakit") or 0)
    rx_banka       = float(e.get("rx_banka")       or 0)
    rx_nakit       = float(e.get("rx_nakit")       or 0)

    banka_top   = teb + vakifbank - iade_banka - tahsilat_banka
    nakit_top   = nakit + iban - iade_nakit - tahsilat_nakit + cikan
    genel_top   = banka_top + nakit_top
    rx_top      = rx_banka + rx_nakit
    fark_banka  = banka_top - rx_banka
    fark_nakit  = nakit_top - rx_nakit
    toplam_fark = fark_banka + fark_nakit

    return dict(
        cikan=cikan, teb=teb, vakifbank=vakifbank, banka_top=banka_top,
        nakit=nakit, iban=iban, iade_nakit=iade_nakit, iade_banka=iade_banka,
        tahsilat_banka=tahsilat_banka, tahsilat_nakit=tahsilat_nakit,
        nakit_top=nakit_top, genel_top=genel_top,
        rx_banka=rx_banka, rx_nakit=rx_nakit, rx_top=rx_top,
        fark_banka=fark_banka, fark_nakit=fark_nakit, toplam_fark=toplam_fark,
        notes=e.get("notes") or "",
    )


def generate_excel(year: int, month: int, entries: dict, holidays: set,
                   hidden_weekdays=None, hidden_dates=None) -> bytes:
    # hidden_weekdays: JS getDay() değerleri (0=Pazar .. 6=Cumartesi)
    hidden_weekdays = set(hidden_weekdays or [])
    hidden_dates    = set(hidden_dates or [])

    wb = Workbook()
    ws = wb.active
    ws.title = f"{TURKISH_MONTHS[month]} {year}"

    # header group fills
    _BANKA  = PatternFill("solid", fgColor="064E3B")
    _BANKA2 = PatternFill("solid", fgColor="065F46")
    _NAKIT  = PatternFill("solid", fgColor="1E3A6E")
    _NAKIT2 = PatternFill("solid", fgColor="1E40AF")
    _IADE   = PatternFill("solid", fgColor="7C2D12")
    _IADE2  = PatternFill("solid", fgColor="9A3412")
    _TAHSIL = PatternFill("solid", fgColor="4C1D95")
    _TAHSIL2= PatternFill("solid", fgColor="6D28D9")
    _RX     = PatternFill("solid", fgColor="1F4E79")
    _FARK   = PatternFill("solid", fgColor="1E3A6E")

    # ── Header row 1 (merged group labels) ────────────────────────────────────
    groups = [
        ("A1:A2", "TARİH",        _H1),
        ("B1:B2", "ÇIKAN",        _H1),
        ("C1:D1", "BANKA (POS)",  _BANKA),
        ("E1:F1", "NAKİT",        _NAKIT),
        ("G1:H1", "İADE",         _IADE),
        ("I1:J1", "TAHSİLAT",     _TAHSIL),
        ("K1:K2", "BANKA TOPLAM", _H1),
        ("L1:L2", "NAKİT TOPLAM", _H1),
        ("M1:M2", "GENEL TOPLAM", _H1),
        ("N1:P1", "RX",           _RX),
        ("Q1:S1", "FARK",         _FARK),
        ("T1:T2", "NOT",          _H1),
    ]
    for rng, label, fill in groups:
        ws.merge_cells(rng)
        top_cell = rng.split(":")[0]
        c = ws[top_cell]
        c.value = label
        c.fill  = fill
        c.font  = _WB
        c.alignment = _CTR
        c.border = _BRD

    # Fill lower half of rowspan-2 single-col cells
    for col_letter in ("A", "B", "K", "L", "M", "T"):
        ws[f"{col_letter}2"].fill = _H1
        ws[f"{col_letter}2"].border = _BRD

    # ── Header row 2 (sub-labels) ──────────────────────────────────────────────
    h2 = [
        ("C", "TEB",           _BANKA2),
        ("D", "VAKİFBANK",     _BANKA2),
        ("E", "GİRİŞ",         _NAKIT2),
        ("F", "İBAN",          _NAKIT2),
        ("G", "BANKA",         _IADE2),
        ("H", "NAKİT",         _IADE2),
        ("I", "BANKA",         _TAHSIL2),
        ("J", "NAKİT",         _TAHSIL2),
        ("N", "BANKA",         _RX),
        ("O", "NAKİT",         _RX),
        ("P", "RX TOPLAM",     _RX),
        ("Q", "BANKA",         _FARK),
        ("R", "NAKİT",         _FARK),
        ("S", "TOPLAM FARK",   _FARK),
    ]
    for col_letter, label, fill in h2:
        c = ws[f"{col_letter}2"]
        c.value = label
        c.fill  = fill
        c.font  = _WB
        c.alignment = _CTR
        c.border = _BRD

    # ── Data rows ──────────────────────────────────────────────────────────────
    num_days = monthrange(year, month)[1]
    totals   = {k: 0.0 for k in [
        "cikan", "teb", "vakifbank", "banka_top",
        "nakit", "iban", "iade_nakit", "iade_banka",
        "tahsilat_banka", "tahsilat_nakit",
        "nakit_top", "genel_top",
        "rx_banka", "rx_nakit", "rx_top",
        "fark_banka", "fark_nakit", "toplam_fark",
    ]}

    row = 3
    for day in range(1, num_days + 1):
        d  = date(year, month, day)
        ds = d.strftime("%Y-%m-%d")

        # Gizli günler tamamen atlanır (toplama da girmez)
        js_day = (d.weekday() + 1) % 7   # Python weekday → JS getDay
        if js_day in hidden_weekdays or ds in hidden_dates:
            continue

        v  = _calc(entries.get(ds, {}))

        for k in totals:
            totals[k] += v.get(k, 0)

        is_weekend = d.weekday() >= 5
        is_holiday = ds in holidays
        bg = _HOL if is_holiday else (_WKD if is_weekend else None)

        day_label = f"{d.day:02d}.{d.month:02d}.{TURKISH_DAYS[d.weekday()]}"

        _c(ws, row,  1, day_label,              fill=bg, align=_CTR)
        _c(ws, row,  2, v["cikan"],             fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row,  3, v["teb"],               fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row,  4, v["vakifbank"],          fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row,  5, v["nakit"],             fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row,  6, v["iban"],              fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row,  7, v["iade_banka"],        fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row,  8, v["iade_nakit"],        fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row,  9, v["tahsilat_banka"],    fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row, 10, v["tahsilat_nakit"],    fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row, 11, v["banka_top"],          fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row, 12, v["nakit_top"],         fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row, 13, v["genel_top"],         fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row, 14, v["rx_banka"],          fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row, 15, v["rx_nakit"],          fill=bg, fmt=_MNY, align=_RGT)
        _c(ws, row, 16, v["rx_top"],            fill=bg, fmt=_MNY, align=_RGT)

        for col, val in [(17, v["fark_banka"]), (18, v["fark_nakit"]), (19, v["toplam_fark"])]:
            if   val < -0.005: fill, font = _RED, Font(color="FFFFFF")
            elif val >  0.005: fill, font = _GRN, Font(color="FFFFFF")
            else:              fill, font = bg,   None
            _c(ws, row, col, val, fill=fill, font=font, fmt=_MNY, align=_RGT)

        _c(ws, row, 20, v["notes"], fill=bg, align=_LFT)
        row += 1

    # ── Monthly total row ──────────────────────────────────────────────────────
    _c(ws, row, 1, "TOPLAM", fill=_TOT, font=_WB)
    total_cols = [
        (2,  "cikan"),          (3,  "teb"),           (4,  "vakifbank"),
        (5,  "nakit"),          (6,  "iban"),
        (7,  "iade_banka"),     (8,  "iade_nakit"),
        (9,  "tahsilat_banka"), (10, "tahsilat_nakit"),
        (11, "banka_top"),      (12, "nakit_top"),     (13, "genel_top"),
        (14, "rx_banka"),       (15, "rx_nakit"),      (16, "rx_top"),
        (17, "fark_banka"),     (18, "fark_nakit"),    (19, "toplam_fark"),
    ]
    for col, key in total_cols:
        _c(ws, row, col, totals[key], fill=_TOT, font=_WB, fmt=_MNY, align=_RGT)
    _c(ws, row, 20, "", fill=_TOT, font=_WB)

    # ── Column widths ──────────────────────────────────────────────────────────
    widths = {
        "A": 13, "B": 12, "C": 12, "D": 12, "E": 12,
        "F": 12, "G": 12, "H": 12, "I": 12,
        "J": 12, "K": 12, "L": 12, "M": 12,
        "N": 12, "O": 12, "P": 12,
        "Q": 12, "R": 12, "S": 12,
        "T": 26,
    }
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "B3"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
