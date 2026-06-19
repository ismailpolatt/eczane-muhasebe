import io
from datetime import datetime

import openpyxl


def get_columns(file_bytes: bytes) -> list[str]:
    """Return column names from the first non-empty row of RxEys Excel."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    headers: list[str] = []
    for row in ws.iter_rows(max_row=5, values_only=True):
        non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 2:
            headers = non_empty
            break
    wb.close()
    return headers


def _parse_date(val) -> str | None:
    """Convert various date representations to 'YYYY-MM-DD'."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if hasattr(val, "strftime"):  # date object
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _to_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("₺", "").replace(" ", "")
    # Handle Turkish format: 1.234,56 → 1234.56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_rxeys(
    file_bytes: bytes,
    date_col: str,
    banka_col: str,
    nakit_col: str,
) -> dict[str, dict]:
    """Parse RxEys Excel and return {date_str: {rx_banka, rx_nakit}}."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Find header row
    header_row_idx = None
    headers: list[str] = []
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        row_strs = [str(c).strip() if c is not None else "" for c in row]
        if date_col in row_strs:
            header_row_idx = i
            headers = row_strs
            break

    if header_row_idx is None:
        wb.close()
        return {}

    try:
        date_idx = headers.index(date_col)
        banka_idx = headers.index(banka_col)
        nakit_idx = headers.index(nakit_col)
    except ValueError:
        wb.close()
        return {}

    result: dict[str, dict] = {}
    for row_i, row in enumerate(ws.iter_rows(values_only=True)):
        if row_i <= header_row_idx:
            continue
        if all(c is None for c in row):
            continue

        date_str = _parse_date(row[date_idx] if date_idx < len(row) else None)
        if not date_str:
            continue

        rx_banka = _to_float(row[banka_idx] if banka_idx < len(row) else None)
        rx_nakit = _to_float(row[nakit_idx] if nakit_idx < len(row) else None)

        result[date_str] = {"rx_banka": rx_banka, "rx_nakit": rx_nakit}

    wb.close()
    return result
